"""
Excel Integration for JAXSR DOE Studies.

Generates experiment template spreadsheets for lab users to fill in,
reads completed templates back with validation, and adds report sheets
with model results and charts.

Requires the ``excel`` optional dependency group::

    pip install jaxsr[excel]

Which installs ``xlsxwriter`` (for template generation) and ``openpyxl``
(for reading completed templates and adding report sheets).
"""

from __future__ import annotations

import hashlib
import json
from typing import TYPE_CHECKING

import numpy as np

if TYPE_CHECKING:
    from .study import DOEStudy


def _check_xlsxwriter():
    """Raise ImportError with install instructions if xlsxwriter is missing."""
    try:
        import xlsxwriter  # noqa: F401
    except ImportError:
        raise ImportError(
            "xlsxwriter is required for Excel template generation. "
            "Install it with: pip install jaxsr[excel]"
        ) from None


def _check_openpyxl():
    """Raise ImportError with install instructions if openpyxl is missing."""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError(
            "openpyxl is required for reading Excel files. "
            "Install it with: pip install jaxsr[excel]"
        ) from None


def _compute_study_fingerprint(study: DOEStudy) -> str:
    """
    Compute a fingerprint for validating Excel templates against their study.

    The fingerprint is a hash of the study's structural identity: name,
    factor names, bounds, feature types, and categories. This detects
    when a template is loaded against the wrong study.

    Parameters
    ----------
    study : DOEStudy
        The study to fingerprint.

    Returns
    -------
    fingerprint : str
        Hex digest of the study's identity hash.
    """
    identity = {
        "name": study.name,
        "factor_names": study.factor_names,
        "bounds": [list(b) for b in study.bounds],
        "feature_types": study.feature_types,
        "categories": (
            {str(k): v for k, v in study.categories.items()} if study.categories else None
        ),
    }
    raw = json.dumps(identity, sort_keys=True).encode()
    return hashlib.sha256(raw).hexdigest()[:16]


# =============================================================================
# Template generation (xlsxwriter)
# =============================================================================


def generate_template(study: DOEStudy, filepath: str) -> str:
    """
    Generate an Excel template for collecting experiment results.

    Creates a ``.xlsx`` file with a **Design** sheet containing pre-filled
    factor columns (locked) and an empty **Response** column (unlocked) for
    the user to fill in. Categorical factors get dropdown validation.
    A hidden **_Metadata** sheet stores the study fingerprint for validation.

    Parameters
    ----------
    study : DOEStudy
        The study containing a design to export.
    filepath : str
        Path for the output ``.xlsx`` file.

    Returns
    -------
    filepath : str
        The path that was written (same as input).

    Raises
    ------
    RuntimeError
        If no design has been created yet.
    ImportError
        If xlsxwriter is not installed.
    """
    _check_xlsxwriter()
    import xlsxwriter

    if study.design_points is None:
        raise RuntimeError("No design created yet. Call study.create_design() first.")

    X_design = study.design_points
    n_runs, n_factors = X_design.shape

    wb = xlsxwriter.Workbook(filepath)

    # --- Formats ---
    header_fmt = wb.add_format(
        {
            "bold": True,
            "bg_color": "#4472C4",
            "font_color": "#FFFFFF",
            "border": 1,
            "text_wrap": True,
            "align": "center",
            "valign": "vcenter",
        }
    )
    locked_fmt = wb.add_format(
        {
            "locked": True,
            "bg_color": "#D9E2F3",
            "border": 1,
            "num_format": "0.000",
            "align": "center",
        }
    )
    locked_text_fmt = wb.add_format(
        {
            "locked": True,
            "bg_color": "#D9E2F3",
            "border": 1,
            "align": "center",
        }
    )
    run_fmt = wb.add_format(
        {
            "locked": True,
            "bg_color": "#D9E2F3",
            "border": 1,
            "align": "center",
            "bold": True,
        }
    )
    response_fmt = wb.add_format(
        {
            "locked": False,
            "bg_color": "#E2EFDA",
            "border": 1,
            "num_format": "0.0000",
            "align": "center",
        }
    )
    notes_fmt = wb.add_format(
        {
            "locked": False,
            "border": 1,
            "text_wrap": True,
        }
    )
    info_header_fmt = wb.add_format({"bold": True, "font_size": 12})
    info_fmt = wb.add_format({"font_size": 11})

    # --- Instructions sheet ---
    ws_info = wb.add_worksheet("Instructions")
    ws_info.set_column("A:A", 80)
    ws_info.protect()
    ws_info.write("A1", f"JAXSR Experiment Template: {study.name}", info_header_fmt)
    ws_info.write("A2", "")

    instructions = [
        "How to use this template:",
        "",
        "1. Go to the 'Design' tab to see your experiment runs.",
        "2. The factor columns (blue) are locked \u2014 do not modify them.",
        "3. Fill in the 'Response' column (green) with your measured values.",
        "4. Optionally add notes in the 'Notes' column.",
        "5. Save the file and load it back with:",
        "       jaxsr add <study_file> <this_file>",
        "   or in Python:",
        "       from jaxsr.excel import read_completed_template",
        "       X, y = read_completed_template(study, 'this_file.xlsx')",
        "",
        f"Study: {study.name}",
        f"Factors: {', '.join(study.factor_names)}",
        f"Design points: {n_runs}",
    ]
    if study.description:
        instructions.append(f"Description: {study.description}")

    for row, line in enumerate(instructions, start=2):
        ws_info.write(row, 0, line, info_fmt)

    # --- Design sheet ---
    ws = wb.add_worksheet("Design")
    ws.protect("", {"objects": True, "scenarios": True})

    # Column widths
    ws.set_column(0, 0, 8)  # Run
    for col in range(n_factors):
        ws.set_column(col + 1, col + 1, max(14, len(study.factor_names[col]) + 4))
    ws.set_column(n_factors + 1, n_factors + 1, 16)  # Response
    ws.set_column(n_factors + 2, n_factors + 2, 30)  # Notes

    # Header row
    ws.write(0, 0, "Run", header_fmt)
    for col, name in enumerate(study.factor_names):
        ws.write(0, col + 1, name, header_fmt)
    ws.write(0, n_factors + 1, "Response", header_fmt)
    ws.write(0, n_factors + 2, "Notes", header_fmt)

    # Freeze header row
    ws.freeze_panes(1, 0)

    # Determine which factors are categorical
    cat_indices = set()
    if study.feature_types:
        cat_indices = {i for i, ft in enumerate(study.feature_types) if ft == "categorical"}

    # Data rows
    for row_idx in range(n_runs):
        ws.write(row_idx + 1, 0, row_idx + 1, run_fmt)
        for col_idx in range(n_factors):
            val = X_design[row_idx, col_idx]
            if col_idx in cat_indices and study.categories and col_idx in study.categories:
                # Write categorical value as string
                cat_val = study.categories[col_idx][int(val)]
                ws.write(row_idx + 1, col_idx + 1, str(cat_val), locked_text_fmt)
            else:
                ws.write(row_idx + 1, col_idx + 1, float(val), locked_fmt)

        # Empty response cell (unlocked)
        ws.write_blank(row_idx + 1, n_factors + 1, None, response_fmt)
        ws.write_blank(row_idx + 1, n_factors + 2, None, notes_fmt)

    # Add dropdown validation for categorical factors
    if study.categories:
        for col_idx, levels in study.categories.items():
            str_levels = [str(lev) for lev in levels]
            ws.data_validation(
                1,
                col_idx + 1,
                n_runs,
                col_idx + 1,
                {
                    "validate": "list",
                    "source": str_levels,
                    "input_title": study.factor_names[col_idx],
                    "input_message": f"Valid levels: {', '.join(str_levels)}",
                },
            )

    # Response column validation (numeric)
    ws.data_validation(
        1,
        n_factors + 1,
        n_runs,
        n_factors + 1,
        {
            "validate": "any",
            "input_title": "Response",
            "input_message": "Enter your measured response value",
        },
    )

    # --- Hidden metadata sheet ---
    ws_meta = wb.add_worksheet("_Metadata")
    ws_meta.hide()
    ws_meta.write("A1", "study_name")
    ws_meta.write("B1", study.name)
    ws_meta.write("A2", "fingerprint")
    ws_meta.write("B2", _compute_study_fingerprint(study))
    ws_meta.write("A3", "n_factors")
    ws_meta.write("B3", n_factors)
    ws_meta.write("A4", "n_runs")
    ws_meta.write("B4", n_runs)
    ws_meta.write("A5", "factor_names")
    ws_meta.write("B5", json.dumps(study.factor_names))
    ws_meta.write("A6", "feature_types")
    ws_meta.write("B6", json.dumps(study.feature_types))
    ws_meta.write("A7", "categories")
    ws_meta.write(
        "B7",
        json.dumps({str(k): v for k, v in study.categories.items()} if study.categories else None),
    )

    wb.close()
    return filepath


# =============================================================================
# Template reading and validation (openpyxl)
# =============================================================================


class TemplateValidationError(ValueError):
    """Raised when a completed template fails validation."""

    pass


def validate_template(study: DOEStudy, filepath: str) -> list[str]:
    """
    Validate a completed Excel template against its study.

    Checks that the template was generated for this study (fingerprint match),
    factor columns are unmodified, and response values are present and numeric.

    Parameters
    ----------
    study : DOEStudy
        The study this template should belong to.
    filepath : str
        Path to the completed ``.xlsx`` file.

    Returns
    -------
    warnings : list of str
        Non-fatal warnings (e.g., some responses missing). An empty list
        means the template is fully valid.

    Raises
    ------
    TemplateValidationError
        If the template is structurally invalid (wrong study, modified
        factor columns, wrong number of columns, etc.).
    ImportError
        If openpyxl is not installed.
    """
    _check_openpyxl()
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    warnings = []

    # Check metadata sheet
    if "_Metadata" not in wb.sheetnames:
        raise TemplateValidationError(
            "This file does not appear to be a JAXSR experiment template "
            "(no _Metadata sheet found)."
        )

    ws_meta = wb["_Metadata"]
    file_fingerprint = ws_meta["B2"].value
    expected_fingerprint = _compute_study_fingerprint(study)

    if file_fingerprint != expected_fingerprint:
        file_study_name = ws_meta["B1"].value or "unknown"
        raise TemplateValidationError(
            f"Template fingerprint mismatch. This template was generated for "
            f"study '{file_study_name}', but you are loading it into "
            f"study '{study.name}'. Use the correct template file."
        )

    # Check Design sheet
    if "Design" not in wb.sheetnames:
        raise TemplateValidationError("No 'Design' sheet found in template.")

    ws = wb["Design"]
    n_runs_expected = study.design_points.shape[0] if study.design_points is not None else 0

    # Check header row
    expected_headers = ["Run"] + study.factor_names + ["Response", "Notes"]
    actual_headers = [cell.value for cell in ws[1][: len(expected_headers)]]
    if actual_headers != expected_headers:
        raise TemplateValidationError(
            f"Column headers don't match. Expected {expected_headers}, " f"got {actual_headers}."
        )

    # Check data rows
    n_responses = 0
    n_blank = 0
    for row_idx in range(2, n_runs_expected + 2):
        response_cell = ws.cell(row=row_idx, column=study.n_factors + 2)
        if response_cell.value is not None and response_cell.value != "":
            try:
                float(response_cell.value)
                n_responses += 1
            except (ValueError, TypeError):
                warnings.append(
                    f"Row {row_idx}: Response value '{response_cell.value}' " f"is not numeric."
                )
        else:
            n_blank += 1

    if n_blank > 0:
        warnings.append(
            f"{n_blank} of {n_runs_expected} response values are blank. "
            f"Partial data will be imported."
        )

    wb.close()
    return warnings


def read_completed_template(
    study: DOEStudy,
    filepath: str,
    skip_validation: bool = False,
) -> tuple[np.ndarray, np.ndarray]:
    """
    Read a completed Excel template and extract observations.

    Validates the template against the study, then reads the factor values
    and response column. Only rows with a valid numeric response are returned.

    Parameters
    ----------
    study : DOEStudy
        The study this template belongs to.
    filepath : str
        Path to the completed ``.xlsx`` file.
    skip_validation : bool
        If True, skip fingerprint validation. Useful for templates that were
        manually edited or created outside of JAXSR.

    Returns
    -------
    X : numpy.ndarray
        Feature matrix of shape ``(n_completed, n_factors)``.
    y : numpy.ndarray
        Response vector of shape ``(n_completed,)``.

    Raises
    ------
    TemplateValidationError
        If validation fails and ``skip_validation`` is False.
    ValueError
        If no valid responses are found.
    ImportError
        If openpyxl is not installed.
    """
    _check_openpyxl()
    import openpyxl

    if not skip_validation:
        warnings = validate_template(study, filepath)
        # Warnings are non-fatal — we proceed with partial data
        for w in warnings:
            import sys

            print(f"Warning: {w}", file=sys.stderr)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["Design"]

    # Determine categorical factor index-to-level mapping
    cat_level_to_idx = {}
    if study.categories:
        for col_idx, levels in study.categories.items():
            cat_level_to_idx[col_idx] = {str(lev): i for i, lev in enumerate(levels)}

    X_rows = []
    y_rows = []
    n_runs = ws.max_row - 1  # exclude header

    for row_idx in range(2, n_runs + 2):
        response_cell = ws.cell(row=row_idx, column=study.n_factors + 2)
        if response_cell.value is None or response_cell.value == "":
            continue
        try:
            y_val = float(response_cell.value)
        except (ValueError, TypeError):
            continue

        # Read factor values
        x_row = []
        for col_idx in range(study.n_factors):
            cell_val = ws.cell(row=row_idx, column=col_idx + 2).value
            if col_idx in cat_level_to_idx:
                # Map categorical string back to numeric index
                idx = cat_level_to_idx[col_idx].get(str(cell_val))
                if idx is not None:
                    x_row.append(float(idx))
                else:
                    x_row.append(float("nan"))
            else:
                x_row.append(float(cell_val))

        X_rows.append(x_row)
        y_rows.append(y_val)

    wb.close()

    if not y_rows:
        raise ValueError("No valid response values found in the template.")

    return np.array(X_rows), np.array(y_rows)


# =============================================================================
# Report sheets (openpyxl)
# =============================================================================


def add_report_sheets(study: DOEStudy, filepath: str) -> str:
    """
    Add analysis report sheets to an existing or new Excel workbook.

    Adds sheets with model results, coefficient analysis, residual analysis,
    and charts. If the file exists, report sheets are appended (existing
    report sheets are replaced). If the file doesn't exist, a new workbook
    is created.

    Parameters
    ----------
    study : DOEStudy
        A fitted study with observations.
    filepath : str
        Path to write the report ``.xlsx`` file.

    Returns
    -------
    filepath : str
        The path that was written.

    Raises
    ------
    RuntimeError
        If the study has no fitted model or no observations.
    ImportError
        If openpyxl is not installed.
    """
    _check_openpyxl()
    import openpyxl
    from openpyxl.chart import BarChart, Reference, ScatterChart
    from openpyxl.styles import Border, Font, PatternFill, Side

    if not study.is_fitted:
        raise RuntimeError("Study has no fitted model. Call study.fit() first.")
    if study.X is None or study.y is None:
        raise RuntimeError("Study has no observations.")

    # Load existing or create new
    try:
        wb = openpyxl.load_workbook(filepath)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Remove existing report sheets to replace them
    report_sheets = ["Summary", "Coefficients", "Predictions", "Residuals"]
    for sheet_name in report_sheets:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    model = study.model
    result = model._result

    # --- Styles ---
    title_font = Font(bold=True, size=14, color="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    metric_fill = PatternFill("solid", fgColor="D9E2F3")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _style_header_row(ws, row, n_cols):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    def _style_data_cell(ws, row, col):
        ws.cell(row=row, column=col).border = thin_border

    # =========================================================================
    # Summary sheet
    # =========================================================================
    ws_summary = wb.create_sheet("Summary")
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 50

    ws_summary.cell(row=1, column=1, value=f"DOE Study Report: {study.name}").font = title_font
    ws_summary.merge_cells("A1:B1")

    summary_rows = [
        ("Study Name", study.name),
        ("Description", study.description or "—"),
        ("Factors", ", ".join(study.factor_names)),
        ("Number of Factors", study.n_factors),
        ("Observations", study.n_observations),
        ("", ""),
        ("Model Expression", result.expression()),
        ("Number of Terms", result.n_terms),
        ("MSE", f"{result.mse:.6g}"),
        ("AIC", f"{result.aic:.4f}"),
        ("BIC", f"{result.bic:.4f}"),
        ("AICc", f"{result.aicc:.4f}"),
        ("", ""),
        ("Design Method", study._design_config.get("method", "—")),
        (
            "Selection Strategy",
            study._model_config.get("strategy", "—") if study._model_config else "—",
        ),
        (
            "Information Criterion",
            study._model_config.get("information_criterion", "—") if study._model_config else "—",
        ),
    ]

    for i, (label, value) in enumerate(summary_rows, start=3):
        cell_a = ws_summary.cell(row=i, column=1, value=label)
        cell_b = ws_summary.cell(row=i, column=2, value=value)
        if label:
            cell_a.font = Font(bold=True)
            cell_a.fill = metric_fill
            cell_a.border = thin_border
            cell_b.border = thin_border

    # =========================================================================
    # Coefficients sheet
    # =========================================================================
    ws_coeff = wb.create_sheet("Coefficients")
    ws_coeff.column_dimensions["A"].width = 35
    ws_coeff.column_dimensions["B"].width = 18
    ws_coeff.column_dimensions["C"].width = 18

    ws_coeff.cell(row=1, column=1, value="Model Coefficients").font = title_font

    headers = ["Basis Function", "Coefficient", "|Coefficient|"]
    for col, h in enumerate(headers, start=1):
        ws_coeff.cell(row=3, column=col, value=h)
    _style_header_row(ws_coeff, 3, len(headers))

    coeffs = np.array(result.coefficients)
    names = result.selected_names

    for i, (name, coeff) in enumerate(zip(names, coeffs, strict=True)):
        row = i + 4
        ws_coeff.cell(row=row, column=1, value=name)
        ws_coeff.cell(row=row, column=2, value=float(coeff))
        ws_coeff.cell(row=row, column=3, value=float(abs(coeff)))
        for col in range(1, 4):
            _style_data_cell(ws_coeff, row, col)

    # Bar chart of coefficient magnitudes
    if len(names) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Coefficient Magnitudes"
        chart.y_axis.title = "|Coefficient|"
        chart.x_axis.title = "Basis Function"
        chart.style = 10

        data_ref = Reference(ws_coeff, min_col=3, min_row=3, max_row=3 + len(names))
        cat_ref = Reference(ws_coeff, min_col=1, min_row=4, max_row=3 + len(names))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.shape = 4
        chart.width = 20
        chart.height = 12
        ws_coeff.add_chart(chart, f"A{5 + len(names)}")

    # =========================================================================
    # Predictions sheet
    # =========================================================================
    ws_pred = wb.create_sheet("Predictions")

    ws_pred.cell(row=1, column=1, value="Predictions vs Actual").font = title_font

    # Compute predictions
    predict_fn = model.to_callable()
    X_np = np.asarray(study.X)
    y_actual = np.asarray(study.y)
    y_pred = predict_fn(X_np)
    residuals = y_actual - y_pred

    # R-squared
    ss_res = float(np.sum(residuals**2))
    ss_tot = float(np.sum((y_actual - np.mean(y_actual)) ** 2))
    r_squared = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0

    ws_pred.cell(row=2, column=1, value=f"R² = {r_squared:.6f}")

    headers = ["Run", "Actual", "Predicted", "Residual", "% Error"]
    for col, h in enumerate(headers, start=1):
        ws_pred.cell(row=4, column=col, value=h)
    _style_header_row(ws_pred, 4, len(headers))

    for col_letter, width in [("A", 8), ("B", 14), ("C", 14), ("D", 14), ("E", 12)]:
        ws_pred.column_dimensions[col_letter].width = width

    for i in range(len(y_actual)):
        row = i + 5
        ws_pred.cell(row=row, column=1, value=i + 1)
        ws_pred.cell(row=row, column=2, value=float(y_actual[i]))
        ws_pred.cell(row=row, column=3, value=float(y_pred[i]))
        ws_pred.cell(row=row, column=4, value=float(residuals[i]))
        pct_err = float(abs(residuals[i] / y_actual[i]) * 100) if y_actual[i] != 0 else float("inf")
        ws_pred.cell(row=row, column=5, value=pct_err)
        for col in range(1, 6):
            _style_data_cell(ws_pred, row, col)

    # Scatter chart: predicted vs actual
    if len(y_actual) > 1:
        chart = ScatterChart()
        chart.title = "Predicted vs Actual"
        chart.x_axis.title = "Actual"
        chart.y_axis.title = "Predicted"
        chart.style = 13
        chart.width = 16
        chart.height = 14

        x_values = Reference(ws_pred, min_col=2, min_row=4, max_row=4 + len(y_actual))
        y_values = Reference(ws_pred, min_col=3, min_row=4, max_row=4 + len(y_actual))
        from openpyxl.chart import Series

        s = Series(y_values, x_values, title="Predictions")
        chart.series.append(s)
        ws_pred.add_chart(chart, "G4")

    # =========================================================================
    # Residuals sheet
    # =========================================================================
    ws_resid = wb.create_sheet("Residuals")
    ws_resid.cell(row=1, column=1, value="Residual Analysis").font = title_font

    headers = ["Run", "Predicted", "Residual"]
    for col, h in enumerate(headers, start=1):
        ws_resid.cell(row=3, column=col, value=h)
    _style_header_row(ws_resid, 3, len(headers))

    for col_letter, width in [("A", 8), ("B", 14), ("C", 14)]:
        ws_resid.column_dimensions[col_letter].width = width

    for i in range(len(y_actual)):
        row = i + 4
        ws_resid.cell(row=row, column=1, value=i + 1)
        ws_resid.cell(row=row, column=2, value=float(y_pred[i]))
        ws_resid.cell(row=row, column=3, value=float(residuals[i]))
        for col in range(1, 4):
            _style_data_cell(ws_resid, row, col)

    # Residuals vs predicted scatter
    if len(y_actual) > 1:
        chart = ScatterChart()
        chart.title = "Residuals vs Predicted"
        chart.x_axis.title = "Predicted"
        chart.y_axis.title = "Residual"
        chart.style = 13
        chart.width = 16
        chart.height = 14

        x_values = Reference(ws_resid, min_col=2, min_row=3, max_row=3 + len(y_actual))
        y_values = Reference(ws_resid, min_col=3, min_row=3, max_row=3 + len(y_actual))
        from openpyxl.chart import Series

        s = Series(y_values, x_values, title="Residuals")
        chart.series.append(s)
        ws_resid.add_chart(chart, "E3")

    wb.save(filepath)
    wb.close()
    return filepath
