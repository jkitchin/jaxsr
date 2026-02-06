"""Tests for Excel template generation, validation, and report sheets."""

from __future__ import annotations

import json

import numpy as np
import pytest

from jaxsr.excel import (
    TemplateValidationError,
    _compute_study_fingerprint,
    add_report_sheets,
    generate_template,
    read_completed_template,
    validate_template,
)
from jaxsr.study import DOEStudy

# =============================================================================
# Fixtures
# =============================================================================


@pytest.fixture
def simple_study():
    """A 2-factor continuous study with a design."""
    study = DOEStudy(
        name="excel_test",
        factor_names=["x1", "x2"],
        bounds=[(0, 1), (0, 1)],
    )
    study.create_design(method="latin_hypercube", n_points=10, random_state=42)
    return study


@pytest.fixture
def categorical_study():
    """A study with mixed continuous/categorical factors."""
    study = DOEStudy(
        name="cat_excel_test",
        factor_names=["temp", "pressure", "catalyst"],
        bounds=[(300, 500), (1, 10), (0, 2)],
        feature_types=["continuous", "continuous", "categorical"],
        categories={2: ["A", "B", "C"]},
    )
    study.create_design(method="latin_hypercube", n_points=8, random_state=42)
    return study


@pytest.fixture
def fitted_study():
    """A fitted study for report testing."""
    np.random.seed(42)
    study = DOEStudy(
        name="fitted_test",
        factor_names=["x1", "x2"],
        bounds=[(0, 1), (0, 1)],
    )
    X = np.random.rand(30, 2)
    y = 2.0 * X[:, 0] + 3.0 * X[:, 1] + 0.01 * np.random.randn(30)
    study.add_observations(X, y, record_iteration=False)
    study.fit(max_terms=3)
    return study


# =============================================================================
# Template generation
# =============================================================================


class TestGenerateTemplate:
    """Test Excel template generation."""

    def test_generates_file(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        result = generate_template(simple_study, filepath)
        assert result == filepath

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        assert "Design" in wb.sheetnames
        assert "Instructions" in wb.sheetnames
        assert "_Metadata" in wb.sheetnames
        wb.close()

    def test_design_sheet_structure(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]

        # Check headers
        assert ws.cell(1, 1).value == "Run"
        assert ws.cell(1, 2).value == "x1"
        assert ws.cell(1, 3).value == "x2"
        assert ws.cell(1, 4).value == "Response"
        assert ws.cell(1, 5).value == "Notes"

        # Check data rows exist
        assert ws.cell(2, 1).value == 1  # Run number
        assert ws.cell(2, 2).value is not None  # x1 value
        assert ws.cell(2, 4).value is None  # Response empty

        # Check all 10 runs
        assert ws.cell(11, 1).value == 10
        assert ws.cell(12, 1).value is None  # no 11th row

        wb.close()

    def test_metadata_sheet(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["_Metadata"]

        assert ws["B1"].value == "excel_test"
        assert ws["B2"].value == _compute_study_fingerprint(simple_study)
        assert ws["B3"].value == 2  # n_factors
        assert ws["B4"].value == 10  # n_runs
        assert json.loads(ws["B5"].value) == ["x1", "x2"]

        wb.close()

    def test_categorical_template(self, categorical_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(categorical_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]

        # Check header
        assert ws.cell(1, 4).value == "catalyst"
        assert ws.cell(1, 5).value == "Response"

        wb.close()

    def test_no_design_raises(self, tmp_path):
        study = DOEStudy(
            name="empty",
            factor_names=["x1"],
            bounds=[(0, 1)],
        )
        with pytest.raises(RuntimeError, match="No design"):
            generate_template(study, str(tmp_path / "bad.xlsx"))


# =============================================================================
# Template validation
# =============================================================================


class TestValidateTemplate:
    """Test template validation logic."""

    def test_valid_template_no_responses(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        warnings = validate_template(simple_study, filepath)
        assert len(warnings) == 1  # All responses blank
        assert "blank" in warnings[0].lower()

    def test_valid_template_with_responses(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        # Fill in all responses
        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]
        for row in range(2, 12):
            ws.cell(row=row, column=4, value=float(row))
        wb.save(filepath)
        wb.close()

        warnings = validate_template(simple_study, filepath)
        assert len(warnings) == 0

    def test_wrong_study_fingerprint(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        other_study = DOEStudy(
            name="different_study",
            factor_names=["a", "b"],
            bounds=[(0, 10), (0, 10)],
        )
        other_study.create_design(method="latin_hypercube", n_points=10, random_state=42)

        with pytest.raises(TemplateValidationError, match="fingerprint mismatch"):
            validate_template(other_study, filepath)

    def test_non_numeric_response_warning(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]
        ws.cell(row=2, column=4, value="not_a_number")
        ws.cell(row=3, column=4, value=42.0)
        wb.save(filepath)
        wb.close()

        warnings = validate_template(simple_study, filepath)
        assert any("not numeric" in w for w in warnings)
        assert any("blank" in w.lower() for w in warnings)

    def test_no_metadata_sheet_raises(self, simple_study, tmp_path):
        filepath = str(tmp_path / "plain.xlsx")
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Design"
        ws.append(["Run", "x1", "x2", "Response"])
        wb.save(filepath)
        wb.close()

        with pytest.raises(TemplateValidationError, match="_Metadata"):
            validate_template(simple_study, filepath)


# =============================================================================
# Reading completed templates
# =============================================================================


class TestReadCompletedTemplate:
    """Test reading completed Excel templates."""

    def test_read_all_responses(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]
        for row in range(2, 12):
            ws.cell(row=row, column=4, value=float(row * 10))
        wb.save(filepath)
        wb.close()

        X, y = read_completed_template(simple_study, filepath)
        assert X.shape == (10, 2)
        assert y.shape == (10,)
        np.testing.assert_array_equal(y, np.arange(2, 12) * 10.0)

    def test_read_partial_responses(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]
        # Only fill in rows 2, 3, 4 (first 3 runs)
        for row in range(2, 5):
            ws.cell(row=row, column=4, value=float(row))
        wb.save(filepath)
        wb.close()

        X, y = read_completed_template(simple_study, filepath)
        assert X.shape == (3, 2)
        assert y.shape == (3,)

    def test_read_no_responses_raises(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        with pytest.raises(ValueError, match="No valid response"):
            read_completed_template(simple_study, filepath)

    def test_skip_validation(self, simple_study, tmp_path):
        filepath = str(tmp_path / "template.xlsx")
        generate_template(simple_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Design"]
        ws.cell(row=2, column=4, value=42.0)
        wb.save(filepath)
        wb.close()

        other_study = DOEStudy(
            name="other",
            factor_names=["x1", "x2"],
            bounds=[(0, 1), (0, 1)],
        )
        other_study.create_design(method="latin_hypercube", n_points=10, random_state=42)

        # Without skip_validation, this would fail
        X, y = read_completed_template(other_study, filepath, skip_validation=True)
        assert len(y) == 1


# =============================================================================
# Fingerprint
# =============================================================================


class TestFingerprint:
    """Test study fingerprinting."""

    def test_same_study_same_fingerprint(self):
        s1 = DOEStudy(name="test", factor_names=["x"], bounds=[(0, 1)])
        s2 = DOEStudy(name="test", factor_names=["x"], bounds=[(0, 1)])
        assert _compute_study_fingerprint(s1) == _compute_study_fingerprint(s2)

    def test_different_name_different_fingerprint(self):
        s1 = DOEStudy(name="test1", factor_names=["x"], bounds=[(0, 1)])
        s2 = DOEStudy(name="test2", factor_names=["x"], bounds=[(0, 1)])
        assert _compute_study_fingerprint(s1) != _compute_study_fingerprint(s2)

    def test_different_factors_different_fingerprint(self):
        s1 = DOEStudy(name="test", factor_names=["x"], bounds=[(0, 1)])
        s2 = DOEStudy(name="test", factor_names=["y"], bounds=[(0, 1)])
        assert _compute_study_fingerprint(s1) != _compute_study_fingerprint(s2)


# =============================================================================
# Report sheets
# =============================================================================


class TestReportSheets:
    """Test Excel report sheet generation."""

    def test_add_report_sheets(self, fitted_study, tmp_path):
        filepath = str(tmp_path / "report.xlsx")
        result = add_report_sheets(fitted_study, filepath)
        assert result == filepath

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        assert "Summary" in wb.sheetnames
        assert "Coefficients" in wb.sheetnames
        assert "Predictions" in wb.sheetnames
        assert "Residuals" in wb.sheetnames
        wb.close()

    def test_summary_sheet_content(self, fitted_study, tmp_path):
        filepath = str(tmp_path / "report.xlsx")
        add_report_sheets(fitted_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Summary"]

        # Check study name is present
        values = [ws.cell(row=r, column=2).value for r in range(1, 20)]
        assert "fitted_test" in values
        wb.close()

    def test_predictions_sheet_content(self, fitted_study, tmp_path):
        filepath = str(tmp_path / "report.xlsx")
        add_report_sheets(fitted_study, filepath)

        import openpyxl

        wb = openpyxl.load_workbook(filepath)
        ws = wb["Predictions"]

        # Check header
        assert ws.cell(4, 1).value == "Run"
        assert ws.cell(4, 2).value == "Actual"
        assert ws.cell(4, 3).value == "Predicted"
        assert ws.cell(4, 4).value == "Residual"

        # Check data rows
        assert ws.cell(5, 1).value == 1
        assert ws.cell(5, 2).value is not None  # Actual value

        wb.close()

    def test_not_fitted_raises(self, tmp_path):
        study = DOEStudy(name="empty", factor_names=["x"], bounds=[(0, 1)])
        with pytest.raises(RuntimeError, match="no fitted model"):
            add_report_sheets(study, str(tmp_path / "bad.xlsx"))

    def test_append_to_existing(self, fitted_study, tmp_path):
        """Report sheets can be added to an existing workbook."""
        filepath = str(tmp_path / "existing.xlsx")

        # Create a workbook with existing content
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MyData"
        ws.cell(1, 1, value="existing data")
        wb.save(filepath)
        wb.close()

        add_report_sheets(fitted_study, filepath)

        wb = openpyxl.load_workbook(filepath)
        assert "MyData" in wb.sheetnames  # Original sheet preserved
        assert "Summary" in wb.sheetnames  # Report added
        wb.close()
